"""Run local document-translation benchmark batches.

This script bypasses the frontend and SaaS upload path. It feeds local files
directly to the translation orchestration layer, saves each translated file with
a unique benchmark name, and appends one row per run to the performance
workbook.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import re
import shutil
import sys
import time
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

from zoneinfo import ZoneInfo


REPO_ROOT = Path(__file__).resolve().parents[2]
TRANSLATION_ROOT = Path(__file__).resolve().parents[1]
if str(TRANSLATION_ROOT) not in sys.path:
    sys.path.insert(0, str(TRANSLATION_ROOT))


DEFAULT_EXCEL_PATH = Path("/Users/yeyeong/gena_genos_serving/성능평가.xlsx")
DEFAULT_OUTPUT_ROOT = TRANSLATION_ROOT / "tmp" / "local_memory_benchmark"

DEFAULT_DOCUMENTS = [
    Path("/Users/yeyeong/Downloads/Public-Procurement-Manual.docx"),
    Path("/Users/yeyeong/Downloads/Twinning_manual_2017_update_2022_EN Corrigendum2026.docx"),
    Path("/Users/yeyeong/Downloads/ECSS-Q-ST-60-13C-Rev.2(30April2025).docx"),
    Path("/Users/yeyeong/Downloads/primary-medical-care-policy-guidance-manual-final.docx"),
    Path("/Users/yeyeong/Downloads/Consultant-Services-Manual-2023.docx"),
    Path("/Users/yeyeong/Downloads/Information Security Manual (March 2023).docx"),
]

DOCUMENT_TYPE_BY_NAME = {
    "Public-Procurement-Manual.docx": "DOCX / 조달 매뉴얼",
    "Twinning_manual_2017_update_2022_EN Corrigendum2026.docx": "DOCX / Twinning 매뉴얼",
    "ECSS-Q-ST-60-13C-Rev.2(30April2025).docx": "DOCX / ECSS 표준 문서",
    "primary-medical-care-policy-guidance-manual-final.docx": "DOCX / 의료 정책 매뉴얼",
    "Consultant-Services-Manual-2023.docx": "DOCX / 서비스 매뉴얼",
    "Information Security Manual (March 2023).docx": "DOCX / 보안 매뉴얼",
}


@dataclass(frozen=True)
class BenchmarkCase:
    repeat_index: int
    document_index: int
    source_path: Path
    memory_enabled: bool


class CaptureLogHandler(logging.Handler):
    def __init__(self) -> None:
        super().__init__(level=logging.INFO)
        self.messages: list[str] = []

    def emit(self, record: logging.LogRecord) -> None:
        self.messages.append(record.getMessage())


@contextmanager
def capture_translation_logs() -> Iterator[CaptureLogHandler]:
    logger = logging.getLogger("uvicorn.error")
    handler = CaptureLogHandler()
    logger.addHandler(handler)
    previous_level = logger.level
    if previous_level > logging.INFO or previous_level == logging.NOTSET:
        logger.setLevel(logging.INFO)
    try:
        yield handler
    finally:
        logger.removeHandler(handler)
        logger.setLevel(previous_level)


def _now_label() -> str:
    try:
        now = datetime.now(ZoneInfo("Asia/Seoul"))
    except Exception:
        now = datetime.now()
    return now.strftime("%y%m%d_%H%M%S")


def _safe_part(value: str, *, limit: int = 120) -> str:
    safe = re.sub(r"[^0-9A-Za-z가-힣_.() -]+", "_", str(value or "").strip())
    safe = re.sub(r"\s+", "_", safe).strip("._- ")
    return safe[:limit] or "document"


def _seconds_label(seconds: float, *, include_minutes: bool = False) -> str:
    if seconds <= 0:
        return "0"
    basic = f"{seconds:.2f}s"
    if not include_minutes:
        return basic
    minutes = int(seconds // 60)
    remain = int(round(seconds - minutes * 60))
    return f"{basic} ({minutes}m {remain:02d}s)"


def _find_float(pattern: str, messages: list[str]) -> float:
    regex = re.compile(pattern)
    for message in messages:
        match = regex.search(message)
        if match:
            try:
                return float(match.group(1))
            except ValueError:
                return 0.0
    return 0.0


def _load_dotenv_files() -> None:
    try:
        from dotenv import load_dotenv
    except Exception:
        return
    load_dotenv(TRANSLATION_ROOT / ".env")
    load_dotenv(TRANSLATION_ROOT / ".env.local", override=True)


def _configure_environment(*, disable_azure: bool) -> None:
    os.environ.setdefault("AI_TRANSLATION_TEMP_GLOSSARY_ENABLED", "1")
    os.environ.setdefault("AI_TRANSLATION_PRE_ANALYSIS_ENABLED", "1")
    os.environ.setdefault("AI_TRANSLATION_INITIAL_TERM_DECISION_ENABLED", "0")
    os.environ.setdefault("AI_TRANSLATION_DOCUMENT_TERM_PRE_JUDGE_ENABLED", "1")
    os.environ.setdefault("AI_TRANSLATION_DOCUMENT_TERM_PRE_JUDGE_ALL_PREFERRED", "1")
    os.environ.setdefault("AI_TRANSLATION_DOCUMENT_TERM_PRE_JUDGE_RAW_CANDIDATES", "1")
    os.environ.setdefault("AI_TRANSLATION_TERM_RESOLVER_ENABLED", "0")
    os.environ.setdefault("AI_TRANSLATION_LLM_CONCURRENCY", "10")
    os.environ.setdefault("AI_TRANSLATION_PROMPT_SNAPSHOT_ENABLED", "1")
    os.environ.setdefault("AI_TRANSLATION_BILINGUAL_SUMMARY_MEMORY_MODE", "external_markdown")
    os.environ.setdefault("AI_TRANSLATION_BILINGUAL_SUMMARY_SELECTIVE_STORE_ENABLED", "1")
    os.environ.setdefault("AI_TRANSLATION_BILINGUAL_SUMMARY_COMPRESSION_ENABLED", "1")
    os.environ.setdefault("AI_TRANSLATION_BILINGUAL_SUMMARY_MARKDOWN_PROMPT_MAX_CHARS", "16000")
    os.environ.setdefault("AI_TRANSLATION_BILINGUAL_SUMMARY_MARKDOWN_BM25_TOP_K", "5")
    if disable_azure:
        os.environ["AZURE_STORAGE_CONNECTION_STRING"] = ""


def _copy_input_for_run(case: BenchmarkCase, run_id: str, output_root: Path) -> Path:
    input_dir = output_root / "inputs"
    input_dir.mkdir(parents=True, exist_ok=True)
    memory_label = "memory-on" if case.memory_enabled else "memory-off"
    stem = _safe_part(case.source_path.stem)
    unique_name = f"{stem}__r{case.repeat_index:02d}__d{case.document_index:02d}__{memory_label}__{run_id}{case.source_path.suffix}"
    destination = input_dir / unique_name
    shutil.copy2(case.source_path, destination)
    return destination


def _copy_translated_result_to_dir(
    result: dict[str, Any],
    *,
    run_id: str,
    source_name: str,
    destination_dir: Path,
) -> str:
    source_file = Path(str(result.get("file_path") or ""))
    if not source_file.exists():
        return ""
    destination_dir.mkdir(parents=True, exist_ok=True)
    output_name = f"{_safe_part(Path(source_name).stem)}__{run_id}__ko{source_file.suffix}"
    destination = destination_dir / output_name
    shutil.copy2(source_file, destination)
    return str(destination)


def _copy_run_files_to_artifact(
    *,
    run_input: Path,
    result: dict[str, Any],
    run_id: str,
    source_name: str,
    metrics: dict[str, Any],
    output_root: Path,
) -> tuple[str, str]:
    artifact_dir = Path(str(metrics.get("artifact_dir") or ""))
    if artifact_dir.exists() and artifact_dir.is_dir():
        uploaded_dir = artifact_dir / "uploaded"
        downloaded_dir = artifact_dir / "downloaded"
    else:
        uploaded_dir = output_root / "inputs"
        downloaded_dir = output_root / "results"

    uploaded_dir.mkdir(parents=True, exist_ok=True)
    uploaded_path = uploaded_dir / run_input.name
    if run_input.resolve() != uploaded_path.resolve():
        shutil.copy2(run_input, uploaded_path)

    translated_path = _copy_translated_result_to_dir(
        result,
        run_id=run_id,
        source_name=source_name,
        destination_dir=downloaded_dir,
    )
    return str(uploaded_path), translated_path


def _artifact_dir_for_job(job_id: str) -> Path | None:
    roots: list[Path] = []
    env_root = os.getenv("AI_TRANSLATION_JOB_ARTIFACT_ROOT", "").strip()
    if env_root:
        env_path = Path(env_root)
        if env_path.is_absolute():
            roots.append(env_path)
        else:
            roots.extend([Path.cwd() / env_path, REPO_ROOT / env_path, TRANSLATION_ROOT / env_path])
    roots.extend(
        [
            REPO_ROOT / "tmp" / "job_artifacts",
            TRANSLATION_ROOT / "tmp" / "job_artifacts",
        ]
    )

    matches: list[Path] = []
    seen: set[Path] = set()
    for root in roots:
        root = root.resolve()
        if root in seen or not root.exists():
            continue
        seen.add(root)
        matches.extend(root.glob(f"*__{job_id}"))
        matches.extend(root.glob(f"*{job_id}*"))
    matches = sorted({path.resolve() for path in matches if path.is_dir()})
    return matches[-1] if matches else None


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _extract_metrics(job_id: str) -> dict[str, Any]:
    artifact_dir = _artifact_dir_for_job(job_id)
    if not artifact_dir:
        return {"artifact_dir": ""}

    summary = _read_json(artifact_dir / "bilingual_summary_memory.json")
    term_memory = _read_json(artifact_dir / "document_term_memory.json")
    pre_judge = _read_json(artifact_dir / "pre_judge.json")
    pre_analysis = _read_json(artifact_dir / "pre_analysis.json")

    pre_judge_actions = ((pre_judge.get("result") or {}).get("actions") or [])
    term_entries = term_memory.get("entries") or {}
    return {
        "artifact_dir": str(artifact_dir),
        "page_count": int(summary.get("scope_count") or 0),
        "source_word_count": int(summary.get("source_word_count") or 0),
        "total_chars": int(summary.get("total_chars") or 0),
        "progress_chars": int(summary.get("total_chars") or 0),
        "memory_store_count": max(
            0,
            int(summary.get("summary_update_call_count") or 0)
            - int(summary.get("summary_update_skip_count") or 0),
        ),
        "memory_skip_count": int(summary.get("summary_update_skip_count") or 0),
        "memory_tokens": int(summary.get("prompt_memory_estimated_tokens") or 0),
        "memory_update_seconds": (int(summary.get("summary_update_total_elapsed_ms") or 0) / 1000.0),
        "term_count": len(term_entries) if isinstance(term_entries, dict) else 0,
        "pre_judge_action_count": len(pre_judge_actions) if isinstance(pre_judge_actions, list) else 0,
        "pre_analysis_exists": bool(pre_analysis),
    }


def _setting_label(memory_enabled: bool) -> str:
    memory = (
        "장문 메모리 ON(선별 저장/압축 ON, BM25 OFF, 프롬프트 길이 제한 OFF)"
        if memory_enabled
        else "장문 메모리 OFF"
    )
    return (
        f"{memory}, 사전분석/용어집/pre-judge ON, initial glossary OFF, "
        "DTM resolver OFF, DOCX 병렬 10, target-language retry ON"
    )


def _last_nonempty_row(ws: Any) -> int:
    for row_index in range(ws.max_row, 0, -1):
        values = [ws.cell(row_index, column).value for column in range(1, ws.max_column + 1)]
        if any(value not in (None, "") for value in values):
            return row_index
    return 1


def _append_excel_row(excel_path: Path, row: list[Any]) -> int:
    from copy import copy

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    sheet = workbook.active
    source_row = _last_nonempty_row(sheet)
    target_row = source_row + 1
    for column_index, value in enumerate(row, start=1):
        cell = sheet.cell(target_row, column_index, value)
        template = sheet.cell(source_row, column_index)
        if template.has_style:
            cell._style = copy(template._style)
        if template.number_format:
            cell.number_format = template.number_format
        if template.alignment:
            cell.alignment = copy(template.alignment)
    workbook.save(excel_path)
    return target_row


async def _run_one_case(
    case: BenchmarkCase,
    *,
    target_lang: str,
    output_root: Path,
    excel_path: Path,
    dry_run: bool,
) -> dict[str, Any]:
    import translation_orchestration

    memory_label = "on" if case.memory_enabled else "off"
    run_id = f"{_now_label()}__r{case.repeat_index:02d}__d{case.document_index:02d}__mem-{memory_label}__{uuid.uuid4().hex[:8]}"

    if dry_run:
        memory_file_label = "memory-on" if case.memory_enabled else "memory-off"
        stem = _safe_part(case.source_path.stem)
        would_copy_to = (
            output_root
            / "inputs"
            / f"{stem}__r{case.repeat_index:02d}__d{case.document_index:02d}__{memory_file_label}__{run_id}{case.source_path.suffix}"
        )
        return {
            "run_id": run_id,
            "source": str(case.source_path),
            "input": str(would_copy_to),
            "memory_enabled": case.memory_enabled,
            "dry_run": True,
        }

    run_input = _copy_input_for_run(case, run_id, output_root)
    filename_for_record = run_input.name

    payload = {
        "file": str(run_input),
        "filename": run_input.name,
        "format": target_lang,
        "is_return_file": True,
        "stream": False,
        "style_options": {
            "_job_id": run_id,
            "_filename": run_input.name,
            "bilingual_summary_memory": case.memory_enabled,
            "bilingual_summary_memory_mode": "external_markdown",
            "bilingual_summary_memory_compression": True,
        },
    }

    started = time.perf_counter()
    with capture_translation_logs() as captured_logs:
        result = await translation_orchestration.run(payload)
    total_seconds = time.perf_counter() - started
    result.pop("file_base64", None)

    messages = captured_logs.messages
    metrics = _extract_metrics(run_id)
    uploaded_path, result_path = _copy_run_files_to_artifact(
        run_input=run_input,
        result=result,
        run_id=run_id,
        source_name=filename_for_record,
        metrics=metrics,
        output_root=output_root,
    )
    original_preview_seconds = _find_float(r"HTML preview URL 생성: ([0-9.]+)s", messages)
    setup_seconds = _find_float(r"Document Term Memory\] setup elapsed ([0-9.]+)s", messages)
    translation_seconds = _find_float(r"LLM 배치 번역: ([0-9.]+)s", messages)
    save_seconds = _find_float(r"파일 저장/다운로드 payload: ([0-9.]+)s", messages)
    memory_update_seconds = metrics.get("memory_update_seconds", 0.0) if case.memory_enabled else 0.0
    memory_status = "ON" if case.memory_enabled else "OFF"
    output_name = Path(result_path).name if result_path else filename_for_record
    page_count = metrics.get("page_count") or ""
    word_char = f"{metrics.get('source_word_count', 0):,} words / {metrics.get('total_chars', 0):,} chars"
    if metrics.get("progress_chars") and metrics.get("progress_chars") != metrics.get("total_chars"):
        word_char = f"{word_char} / progress chars {int(metrics['progress_chars']):,}"

    row = [
        output_name,
        DOCUMENT_TYPE_BY_NAME.get(case.source_path.name, "DOCX / 문서"),
        page_count,
        word_char,
        "English",
        target_lang,
        _setting_label(case.memory_enabled),
        _seconds_label(total_seconds, include_minutes=True),
        _seconds_label(original_preview_seconds),
        _seconds_label(setup_seconds),
        _seconds_label(translation_seconds),
        _seconds_label(memory_update_seconds) if case.memory_enabled else 0,
        _seconds_label(save_seconds),
        0,
        memory_status,
        metrics.get("memory_store_count", 0) if case.memory_enabled else 0,
        metrics.get("memory_skip_count", 0) if case.memory_enabled else 0,
        metrics.get("memory_tokens", 0) if case.memory_enabled else 0,
        "미검수",
        (
            f"job_id={run_id}. DTM {metrics.get('term_count', 0)} entries. "
            f"pre-judge {metrics.get('pre_judge_action_count', 0)} applied. "
            f"artifact={metrics.get('artifact_dir', '')}"
        ),
    ]
    excel_row = _append_excel_row(excel_path, row)

    summary = {
        "run_id": run_id,
        "repeat_index": case.repeat_index,
        "document_index": case.document_index,
        "source_path": str(case.source_path),
        "input_path": str(run_input),
        "uploaded_path": uploaded_path,
        "translated_path": result_path,
        "excel_path": str(excel_path),
        "excel_row": excel_row,
        "memory_enabled": case.memory_enabled,
        "total_seconds": total_seconds,
        "metrics": metrics,
        "translation_status": result.get("translation_status"),
        "translation_error": result.get("translation_error"),
    }
    output_root.mkdir(parents=True, exist_ok=True)
    with (output_root / "benchmark_runs.jsonl").open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(summary, ensure_ascii=False) + "\n")
    return summary


def _build_cases(paths: list[Path], repeats: int) -> list[BenchmarkCase]:
    cases: list[BenchmarkCase] = []
    for repeat_index in range(1, repeats + 1):
        for document_index, path in enumerate(paths, start=1):
            cases.append(BenchmarkCase(repeat_index, document_index, path, False))
            cases.append(BenchmarkCase(repeat_index, document_index, path, True))
    return cases


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run local translation memory benchmark.")
    parser.add_argument("--repeats", type=int, default=10)
    parser.add_argument("--target-lang", default="Korean")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--start-index", type=int, default=1, help="1-based case index to start from.")
    parser.add_argument("--limit", type=int, default=0, help="Maximum number of cases to run. 0 means all.")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--keep-azure", action="store_true", help="Do not clear AZURE_STORAGE_CONNECTION_STRING.")
    parser.add_argument("--documents", nargs="*", type=Path, default=DEFAULT_DOCUMENTS)
    return parser.parse_args()


async def _async_main() -> int:
    args = _parse_args()
    _load_dotenv_files()
    _configure_environment(disable_azure=not args.keep_azure)

    missing = [str(path) for path in args.documents if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing benchmark documents: " + ", ".join(missing))
    if not args.excel.exists() and not args.dry_run:
        raise FileNotFoundError(f"Performance workbook not found: {args.excel}")

    cases = _build_cases(args.documents, args.repeats)
    start = max(1, int(args.start_index)) - 1
    selected = cases[start:]
    if args.limit > 0:
        selected = selected[: args.limit]

    print(
        f"benchmark cases={len(selected)} total_cases={len(cases)} "
        f"start_index={start + 1} dry_run={args.dry_run}"
    )
    for offset, case in enumerate(selected, start=start + 1):
        label = "ON" if case.memory_enabled else "OFF"
        print(f"[{offset}/{len(cases)}] repeat={case.repeat_index} doc={case.document_index} memory={label} {case.source_path.name}", flush=True)
        summary = await _run_one_case(
            case,
            target_lang=args.target_lang,
            output_root=args.output_root,
            excel_path=args.excel,
            dry_run=args.dry_run,
        )
        print(json.dumps(summary, ensure_ascii=False), flush=True)
    return 0


def main() -> int:
    return asyncio.run(_async_main())


if __name__ == "__main__":
    raise SystemExit(main())
