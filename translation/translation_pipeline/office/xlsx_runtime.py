"""XLSX 추출/주입 런타임."""

from __future__ import annotations

import re
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.utils.cell import range_boundaries

from translation_pipeline.common.logging_utils import log_info
from translation_pipeline.common.nodes import is_translatable

from .runtime_common import _element_type_with_placeholder, _node_translation, _preview_bbox
from .xlsx_formatting import _xlsx_display_text


def _xlsx_column_units(sheet: Any, col_index: int) -> float:
    width = getattr(sheet.column_dimensions.get(get_column_letter(col_index)), "width", None)
    if width is None:
        width = 8.43
    return max(1.0, float(width))


def _xlsx_row_units(sheet: Any, row_index: int) -> float:
    height = getattr(sheet.row_dimensions.get(row_index), "height", None)
    if height is None:
        height = 15.0
    return max(1.0, float(height))


def _xlsx_merge_bounds(sheet: Any) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
    merged: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        bounds = (min_row, min_col, max_row, max_col)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged[(row, col)] = bounds
    return merged


def _xlsx_cell_bbox(
    sheet: Any,
    row: int,
    col: int,
    row_span: int = 1,
    col_span: int = 1,
) -> List[int]:
    base_cell_w = 108.0
    base_cell_h = 32.0
    default_col_units = 8.43
    default_row_units = 15.0

    x_units = sum(_xlsx_column_units(sheet, index) for index in range(1, max(1, col)))
    y_units = sum(_xlsx_row_units(sheet, index) for index in range(1, max(1, row)))
    w_units = sum(_xlsx_column_units(sheet, index) for index in range(col, col + max(1, col_span)))
    h_units = sum(_xlsx_row_units(sheet, index) for index in range(row, row + max(1, row_span)))

    x = 62 + x_units / default_col_units * base_cell_w
    y = 58 + y_units / default_row_units * base_cell_h
    width = w_units / default_col_units * base_cell_w
    height = h_units / default_row_units * base_cell_h
    return _preview_bbox(x, y, width, height)



def extract_xlsx(file_path: str) -> Tuple[Any, List[dict]]:
    """XLSX 문서에서 번역 가능한 셀 노드를 추출한다.

    Args:
        file_path: 입력 XLSX 경로.

    Returns:
        워크북 객체와 노드 목록.
    """

    workbook = load_workbook(file_path, data_only=False)
    setattr(workbook, "_ai_translation_source_path", file_path)
    cached_workbook = load_workbook(file_path, data_only=True)
    cached_sheets = {sheet.title: sheet for sheet in cached_workbook.worksheets}
    nodes: List[dict] = []
    for sheet in workbook.worksheets:
        cached_sheet = cached_sheets.get(sheet.title)
        merged_bounds = _xlsx_merge_bounds(sheet)
        nonempty_rows = []
        nonempty_cols = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    nonempty_rows.append(cell.row)
                    nonempty_cols.append(cell.column)
        first_nonempty_row = min(nonempty_rows) if nonempty_rows else None
        first_nonempty_col = min(nonempty_cols) if nonempty_cols else None
        for row in sheet.iter_rows():
            for cell in row:
                raw_value = cell.value
                if isinstance(raw_value, str) and raw_value.startswith("="):
                    continue
                cached_cell = cached_sheet[cell.coordinate] if cached_sheet is not None else None
                value = _xlsx_display_text(cell, cached_cell)
                if is_translatable(value):
                    min_row, min_col, max_row, max_col = merged_bounds.get(
                        (cell.row, cell.column),
                        (cell.row, cell.column, cell.row, cell.column),
                    )
                    row_span = max_row - min_row + 1
                    col_span = max_col - min_col + 1
                    if cell.row == first_nonempty_row:
                        element_type = "column_header"
                    elif cell.column == first_nonempty_col:
                        element_type = "row_header"
                    else:
                        element_type = "table_cell"
                    element_type = _element_type_with_placeholder(value, element_type)
                    nodes.append(
                        {
                            "type": "cell",
                            "group": "sheet_cell",
                            "doc_format": "xlsx",
                            "element_type": element_type,
                            "cell": cell,
                            "text": value,
                            "sheet_name": sheet.title,
                            "row": cell.row,
                            "col": cell.column,
                            "row_index": cell.row,
                            "col_index": cell.column,
                            "cell_ref": cell.coordinate,
                            "row_span": row_span,
                            "col_span": col_span,
                            "is_header": element_type in {"column_header", "row_header"},
                            "merged_range": (
                                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                                if row_span > 1 or col_span > 1
                                else ""
                            ),
                            "bbox": _xlsx_cell_bbox(sheet, min_row, min_col, row_span, col_span),
                        }
                    )
    log_info(f"[XLSX 추출] {len(nodes)}개 번역 노드 추출 완료")
    return workbook, nodes



def inject_xlsx(workbook: Any, nodes: List[dict], trans_map: Dict[str, str]) -> None:
    """XLSX 셀에 번역 결과를 주입한다.

    Args:
        workbook: 워크북 객체.
        nodes: 셀 노드 목록.
        trans_map: 원문/번역 매핑.

    Returns:
        없음.
    """

    count = 0
    sheet_rename_count = 0
    sheet_rename_map: Dict[str, str] = {}
    existing_sheet_names = set(getattr(workbook, "sheetnames", []) or [])

    # 1. 시트 이름을 먼저 확정한다. 이후 셀 수식 참조와 셀 번역은 확정된 시트명 기준으로 처리한다.
    for node in nodes:
        if node.get("type") != "sheet_name":
            continue
        original = str(node["text"])
        translated = _node_translation(node, trans_map)
        if not translated or translated == original:
            continue

        sheet = node.get("sheet")
        if sheet is None:
            continue
        original_title = str(getattr(sheet, "title", "") or node.get("sheet_name", ""))
        new_title = _unique_xlsx_sheet_title(translated, original_title, existing_sheet_names)
        if new_title and new_title != original_title:
            existing_sheet_names.discard(original_title)
            sheet.title = new_title
            existing_sheet_names.add(new_title)
            sheet_rename_map[original_title] = new_title
            sheet_rename_count += 1

    # 2. 시트명 참조가 들어간 수식/이름 정의를 먼저 갱신한다.
    if sheet_rename_map:
        _rewrite_xlsx_sheet_references(workbook, sheet_rename_map)

    # 3. 수식이 아닌 일반 셀 번역을 적용한다.
    for node in nodes:
        if node.get("type") == "sheet_name":
            continue
        original = str(node["text"])
        translated = _node_translation(node, trans_map)
        if not translated or translated == original:
            continue

        cell = node["cell"]
        cell.value = translated
        count += 1
    log_info(f"[XLSX 주입] {count}개 셀 번역 적용, {sheet_rename_count}개 시트명 변경")


def _unique_xlsx_sheet_title(
    title: str,
    current_title: str,
    existing_sheet_names: set[str],
) -> str:
    """Excel sheet title 제약을 맞추고 중복을 피한다."""

    cleaned = re.sub(r"[\[\]:*?/\\]", " ", title).strip() or current_title
    cleaned = re.sub(r"\s+", " ", cleaned)
    base = cleaned[:31] or current_title[:31]
    if base == current_title or base not in existing_sheet_names:
        return base

    for index in range(2, 100):
        suffix = f" ({index})"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in existing_sheet_names:
            return candidate
    return base[:28] + "..."


def _rewrite_xlsx_sheet_references(workbook: Any, rename_map: Dict[str, str]) -> None:
    """시트명 변경 후 수식/이름 정의 안의 시트 참조를 함께 갱신한다."""

    def replace_reference_text(value: str) -> str:
        updated = value
        for old_name, new_name in rename_map.items():
            old_quoted = quote_sheetname(old_name)
            new_quoted = quote_sheetname(new_name)
            updated = updated.replace(f"{old_quoted}!", f"{new_quoted}!")
            updated = re.sub(
                rf"(?<![\w\]']){re.escape(old_name)}(?=!)",
                new_quoted,
                updated,
            )
        return updated

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    cell.value = replace_reference_text(value)

        data_validations = getattr(getattr(worksheet, "data_validations", None), "dataValidation", None)
        if data_validations:
            for validation in data_validations:
                for attr in ("formula1", "formula2"):
                    value = getattr(validation, attr, None)
                    if isinstance(value, str):
                        setattr(validation, attr, replace_reference_text(value))

    defined_names = getattr(workbook, "defined_names", None)
    values_fn = getattr(defined_names, "values", None)
    if callable(values_fn):
        for defined_name in values_fn():
            text = getattr(defined_name, "attr_text", None)
            if isinstance(text, str):
                defined_name.attr_text = replace_reference_text(text)
