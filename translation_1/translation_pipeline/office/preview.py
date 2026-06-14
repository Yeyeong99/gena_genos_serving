"""Office 문서 preview 단계 모듈."""

from __future__ import annotations

from translation_pipeline.common.logging_utils import log_info

import logging
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import time
import tempfile
from html import escape, unescape
from urllib.parse import urlsplit, urlunsplit
import uuid

_logger = logging.getLogger("uvicorn.error")


def _plain_info(message: str, *args: object) -> None:
    log_info(message % args if args else message)

from translation_pipeline.common.azure_uploader import (
    is_azure_preview_enabled,
    upload_html_assets_to_azure,
    upload_html_to_azure,
)
from translation_pipeline.common.preview import (
    _convert_office_to_pdf,
    _render_pdf_preview_svgs,
    cleanup_preview_output_dir,
)

from .types import OfficePipelineDeps, PreviewPayload


def _ensure_aspose_runtime() -> None:
    """Aspose HTML export에 필요한 macOS runtime 경로를 보정한다."""

    repo_root = Path(__file__).resolve().parents[2]
    shim_path = repo_root / "liblibgdiplus.dylib"
    for candidate in ("/opt/homebrew/lib/libgdiplus.dylib", "/usr/local/lib/libgdiplus.dylib"):
        if os.path.exists(candidate):
            if not shim_path.exists():
                try:
                    shim_path.symlink_to(candidate)
                except FileExistsError:
                    pass
            break

    fallback_paths = []
    for path in (str(repo_root), "/opt/homebrew/lib", "/usr/local/lib"):
        if path and os.path.exists(path):
            fallback_paths.append(path)
    existing = os.environ.get("DYLD_FALLBACK_LIBRARY_PATH", "")
    merged = [item for item in existing.split(":") if item]
    for path in fallback_paths:
        if path not in merged:
            merged.insert(0, path)
    if merged:
        os.environ["DYLD_FALLBACK_LIBRARY_PATH"] = ":".join(merged)


def _get_aspose_license_path() -> str | None:
    """Aspose.Total 라이선스 파일 경로를 환경변수에서 읽는다."""

    license_path = os.getenv("AI_TRANSLATION_ASPOSE_LICENSE_PATH") or os.getenv("ASPOSE_LICENSE_PATH")
    if not license_path:
        return None
    expanded = os.path.expanduser(license_path)
    return expanded if os.path.exists(expanded) else None


def _apply_aspose_slides_license(slides_module) -> None:
    """Aspose.Slides 프로세스에 라이선스를 적용한다."""

    license_path = _get_aspose_license_path()
    if not license_path:
        return
    try:
        license = slides_module.License()
        license.set_license(license_path)
    except Exception as exc:
        log_info(f"[Aspose License] Slides 라이선스 적용 실패: {exc}")


_HTML_ASSET_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg", ".bmp", ".css", ".js", ".woff", ".woff2", ".ttf", ".otf"}


def _collect_html_assets(html_path: Path) -> list[Path]:
    """HTML 파일과 같은 디렉터리에 있는 보조 자산 (LibreOffice 가 떨어뜨린 이미지·CSS·폰트) 을 수집한다."""

    output_dir = html_path.parent
    if not output_dir.is_dir():
        return []
    assets: list[Path] = []
    for entry in output_dir.iterdir():
        if not entry.is_file():
            continue
        if entry == html_path:
            continue
        if entry.suffix.lower() in _HTML_ASSET_SUFFIXES:
            assets.append(entry)
    return assets


_HTML_ATTR_REF_RE = re.compile(
    r"""(?P<prefix>\b(?:src|href)\s*=\s*)(?P<quote>["'])(?P<value>[^"'>]+)(?P=quote)""",
    re.IGNORECASE,
)
_CSS_URL_REF_RE = re.compile(
    r"""url\(\s*(?P<quote>["']?)(?P<value>[^"'()\s]+)(?P=quote)\s*\)""",
    re.IGNORECASE,
)


def _resolve_relative_asset(value: str, asset_url_map: dict[str, str]) -> str | None:
    """``./foo.gif`` / 단순 파일명 / 공백 포함 경로를 매핑에서 SAS URL 로 해석한다."""

    name = value.strip()
    if name.startswith("./"):
        name = name[2:]
    return asset_url_map.get(name)


def _rewrite_html_asset_refs(html_path: Path, asset_url_map: dict[str, str]) -> None:
    """HTML 의 ``src``/``href``/CSS ``url(...)`` 참조를 절대 SAS URL 로 치환한다.

    LibreOffice 가 docx→html 로 변환할 때 ``<img src="tmp_html_xxx.gif">`` 또는
    ``<img src="./tmp_html_xxx.gif">`` 처럼 상대 경로 참조를 발행하는데, iframe 이
    cross-origin (Azure Blob) 으로 HTML 을 로드하면 base 가 Azure 도메인이라도
    SAS 쿼리가 없어 403 으로 실패한다. 따라서 HTML 안의 상대 참조를 매핑된 절대
    SAS URL 로 직접 치환한다 — ``src=``/``href=`` 속성과 CSS ``url(...)`` 모두 지원.

    Args:
        html_path: 수정할 로컬 HTML 파일 경로.
        asset_url_map: ``{filename: absolute_sas_url}`` 매핑.
    """

    if not asset_url_map:
        return

    try:
        html = html_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return

    def _attr_repl(match: re.Match[str]) -> str:
        resolved = _resolve_relative_asset(match.group("value"), asset_url_map)
        if resolved is None:
            return match.group(0)
        quote = match.group("quote")
        return f"{match.group('prefix')}{quote}{resolved}{quote}"

    def _url_repl(match: re.Match[str]) -> str:
        resolved = _resolve_relative_asset(match.group("value"), asset_url_map)
        if resolved is None:
            return match.group(0)
        quote = match.group("quote")
        return f"url({quote}{resolved}{quote})"

    new_html, attr_count = _HTML_ATTR_REF_RE.subn(_attr_repl, html)
    new_html, url_count = _CSS_URL_REF_RE.subn(_url_repl, new_html)

    if attr_count + url_count == 0:
        return

    try:
        html_path.write_text(new_html, encoding="utf-8")
    except OSError as exc:
        _logger.warning("[office preview] HTML asset 재작성 저장 실패 (%s): %s", html_path, exc)


def _upload_html_with_assets(
    html_path: Path,
    *,
    job_token: str,
    subdir: str,
) -> str | None:
    """HTML 옆 보조 자산을 먼저 Azure 에 업로드하고 HTML 의 상대 참조를 절대 SAS URL 로
    치환한 뒤 HTML 자체를 업로드해 SAS URL 을 반환한다."""

    assets = _collect_html_assets(html_path)
    if assets:
        asset_url_map = upload_html_assets_to_azure(
            assets,
            job_token=job_token,
            subdir=subdir,
        )
        _rewrite_html_asset_refs(html_path, asset_url_map)
    return upload_html_to_azure(
        html_path,
        job_token=job_token,
        subdir=subdir,
    )


def build_pptx_html_preview_url(
    file_path: str,
    preview_output_dir: str,
    preview_base_url: str = "",
    *,
    job_token: str | None = None,
    subdir: str = "libreoffice-svg-html",
    visible_slides: int | None = None,
) -> str | None:
    """LibreOffice PDF 렌더 SVG 를 인라인한 PPTX preview HTML URL 을 반환한다.

    HTML 단일 파일에 슬라이드 SVG 가 인라인되어 Azure Blob 에 업로드되며 SAS URL
    을 반환한다 (보조 자산 별도 업로드 불필요). ``preview_base_url`` 은 더 이상
    사용하지 않으며 — 정적 서빙은 제거됐다 — Azure 가 비활성이면 ``None`` 을
    반환한다.
    """

    _plain_info(
        "[pptx-preview] build start — file=%s subdir=%s job_token=%s visible_slides=%s preview_dir=%s",
        file_path,
        subdir,
        job_token,
        visible_slides,
        preview_output_dir,
    )

    if not preview_output_dir:
        _logger.warning(
            "[pptx-preview] build skip — preview_output_dir 미설정 (AI_TRANSLATION_PREVIEW_ROOT 확인 필요)"
        )
        return None
    if not is_azure_preview_enabled():
        _logger.warning(
            "[pptx-preview] build skip — Azure preview 비활성 (AZURE_STORAGE_CONNECTION_STRING 빈 값)"
        )
        return None

    try:
        if not job_token:
            cleanup_preview_output_dir(preview_output_dir)
        job_id = job_token or uuid.uuid4().hex
        job_dir = os.path.join(preview_output_dir, job_id, subdir)
        os.makedirs(job_dir, exist_ok=True)
        html_path = os.path.join(job_dir, "index.html")
        _export_pptx_svg_html(
            file_path,
            html_path,
            visible_slides=visible_slides,
        )
        url = upload_html_to_azure(
            Path(html_path),
            job_token=job_id,
            subdir=subdir,
        )
        if url is None:
            _logger.warning(
                "[pptx-preview] build 실패 — upload_html_to_azure 가 None 반환 (위 azure_uploader 로그 확인). html=%s",
                html_path,
            )
        else:
            _plain_info(
                "[pptx-preview] build 성공 — job_id=%s subdir=%s html=%s",
                job_id,
                subdir,
                html_path,
            )
        return url
    except Exception as exc:
        _logger.exception(
            "[pptx-preview] build 실패 — file=%s subdir=%s: %s",
            file_path,
            subdir,
            exc,
        )
        return None


_SVG_ID_ATTR_RE = re.compile(r'(\bid\s*=\s*["\'])([^"\']+)(["\'])', re.IGNORECASE)
# url(#foo), xlink:href="#foo", href="#foo" 세 형태의 fragment 참조를 모두 잡는다.
# 속성명·`HREF` 같은 케이스 변형과 숫자로 시작하는 id 도 누락 없이 격리하도록 IGNORECASE +
# 관대한 char class(`"`/`'`/`)` 가 아닌 모든 문자) 로 컴파일한다.
_SVG_ID_REF_RE = re.compile(
    r'(url\(\s*#|(?:xlink:)?href\s*=\s*["\']\s*#)([^"\')]+)',
    re.IGNORECASE,
)


def _isolate_svg_ids(svg: str, prefix: str) -> str:
    """단일 HTML 안에 여러 SVG 를 인라인할 때 ``id`` 충돌을 막기 위해 prefix 를 부여한다.

    PyMuPDF 는 페이지마다 ``clip_1`` 같은 동일 id 를 발행해 그대로 합치면 두 번째
    슬라이드의 ``url(#clip_1)`` 가 첫 슬라이드 ``<defs>`` 를 가리켜 렌더링이 깨진다.
    HTML 문서 전역에서 id 가 unique 하도록 슬라이드 단위 prefix 를 강제한다.
    """

    def _attr_repl(match: re.Match[str]) -> str:
        return f"{match.group(1)}{prefix}{match.group(2)}{match.group(3)}"

    def _ref_repl(match: re.Match[str]) -> str:
        return f"{match.group(1)}{prefix}{match.group(2)}"

    svg = _SVG_ID_ATTR_RE.sub(_attr_repl, svg)
    svg = _SVG_ID_REF_RE.sub(_ref_repl, svg)
    return svg


def _strip_svg_xml_decl(svg: str) -> str:
    """``<?xml ...?>`` 선언 제거 — HTML 안에 인라인 임베드 시 파서 충돌을 방지한다."""

    stripped = svg.lstrip()
    if stripped.startswith("<?xml"):
        end = stripped.find("?>")
        if end >= 0:
            return stripped[end + 2 :].lstrip()
    return svg


def _export_pptx_svg_html(
    file_path: str,
    html_path: str,
    *,
    visible_slides: int | None = None,
) -> None:
    """PPTX 를 PDF 로 변환한 뒤 각 페이지를 SVG 로 추출해 인라인 HTML 로 저장한다.

    이전 PNG 버전은 텍스트가 픽셀에 박혀 (a) 텍스트 선택·검색·복사 불가 (b) 줌 시
    픽셀화 (c) 향후 블록 단위 편집·하이라이트 토대 부재 라는 한계가 있었다. SVG
    인라인은 ``<text>`` 가 살아남아 위 세 가지를 모두 해소한다 — 슬라이드 단위 id
    prefix 로 충돌 방지.

    Args:
        file_path: 입력 PPTX 경로.
        html_path: 저장할 HTML 경로.
        visible_slides: 1-base 로 노출할 슬라이드 수 — 점진 표시용.
    """

    output_path = Path(html_path)
    output_dir = output_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    # 같은 디렉터리에 남아 있을 수 있는 구버전 (PNG/SVG/PDF) 산출물을 정리해 캐시
    # 충돌을 막는다. 신규 출력은 인라인 HTML 단일 파일이므로 보조 자산이 없다.
    for stale in (
        list(output_dir.glob("slide-*.png"))
        + list(output_dir.glob("slide-*.svg"))
        + list(output_dir.glob("*.pdf"))
    ):
        try:
            stale.unlink()
        except OSError:
            pass

    with tempfile.TemporaryDirectory(prefix="pptx-svg-preview-") as tmpdir:
        pdf_path = _convert_office_to_pdf(file_path, tmpdir)
        svgs = _render_pdf_preview_svgs(pdf_path)

    if isinstance(visible_slides, int) and visible_slides > 0:
        svgs = svgs[:visible_slides]

    slide_sections: list[str] = []
    for index, svg in enumerate(svgs, start=1):
        cleaned = _strip_svg_xml_decl(svg)
        isolated = _isolate_svg_ids(cleaned, f"s{index}_")
        slide_sections.append(
            f'<section class="slide" data-slide="{index}">{isolated}</section>'
        )

    html = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    html, body {{
      margin: 0;
      min-height: 100%;
      background: #f7f9ff;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }}
    body {{
      padding: 24px;
      box-sizing: border-box;
    }}
    .deck {{
      display: flex;
      flex-direction: column;
      gap: 24px;
      align-items: center;
    }}
    .slide {{
      width: min(100%, 1180px);
      margin: 0;
      background: #fff;
      border: 1px solid #dfe7f4;
      box-shadow: 0 14px 40px rgba(15, 23, 42, 0.10);
      overflow: hidden;
    }}
    .slide svg {{
      display: block;
      width: 100%;
      height: auto;
    }}
  </style>
</head>
<body>
  <main class="deck">
    {''.join(slide_sections)}
  </main>
</body>
</html>
"""
    output_path.write_text(html, encoding="utf-8")


def build_docx_html_preview_url(
    file_path: str,
    preview_output_dir: str,
    preview_base_url: str = "",
    *,
    job_token: str | None = None,
    subdir: str = "libreoffice-html",
) -> str | None:
    """LibreOffice를 이용해 DOCX 원본 preview HTML을 생성하고 Azure SAS URL을 반환한다.

    정적 서빙은 제거됐다 — Azure 비활성 시 ``None`` 반환.
    """

    _plain_info(
        "[docx-preview] build start — file=%s subdir=%s job_token=%s preview_dir=%s",
        file_path,
        subdir,
        job_token,
        preview_output_dir,
    )

    if not preview_output_dir:
        _logger.warning(
            "[docx-preview] build skip — preview_output_dir 미설정 (AI_TRANSLATION_PREVIEW_ROOT 확인 필요)"
        )
        return None
    if not is_azure_preview_enabled():
        _logger.warning(
            "[docx-preview] build skip — Azure preview 비활성 (AZURE_STORAGE_CONNECTION_STRING 빈 값)"
        )
        return None

    try:
        if not job_token:
            cleanup_preview_output_dir(preview_output_dir)

        job_id = job_token or uuid.uuid4().hex
        job_dir = os.path.join(preview_output_dir, job_id, subdir)
        os.makedirs(job_dir, exist_ok=True)
        html_path = os.path.join(job_dir, "index.html")

        _export_office_html_with_libreoffice(file_path, html_path)
        # LibreOffice 가 HTML 옆에 ``tmp*_html_xxx.gif`` 같은 보조 이미지를 떨어뜨리므로
        # 같이 Azure 에 업로드하고 HTML 의 상대 참조를 절대 SAS URL 로 재작성한다.
        url = _upload_html_with_assets(
            Path(html_path),
            job_token=job_id,
            subdir=subdir,
        )
        if url is None:
            _logger.warning(
                "[docx-preview] build 실패 — _upload_html_with_assets 가 None 반환. html=%s",
                html_path,
            )
        else:
            _plain_info(
                "[docx-preview] build 성공 — job_id=%s subdir=%s html=%s",
                job_id,
                subdir,
                html_path,
            )
        return url
    except Exception as exc:
        _logger.exception(
            "[docx-preview] build 실패 — file=%s subdir=%s: %s",
            file_path,
            subdir,
            exc,
        )
        return None


def _find_libreoffice_bin() -> str:
    candidates = [
        os.getenv("LIBREOFFICE_BIN", ""),
        shutil.which("soffice") or "",
        shutil.which("libreoffice") or "",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
    ]
    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return ""


def _export_office_html_with_libreoffice(file_path: str, html_path: str) -> None:
    """LibreOffice headless로 Office 문서를 HTML로 직접 변환한다."""

    libreoffice_bin = _find_libreoffice_bin()
    if not libreoffice_bin:
        raise RuntimeError("LibreOffice 실행 파일을 찾을 수 없음")

    output_path = Path(html_path)
    output_dir = output_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    # HTML 과 LibreOffice 가 떨어뜨린 보조 자산을 함께 정리해 직전 실행분이 다음 변환에
    # 잔존해 잘못된 SAS URL 매핑을 만드는 회귀를 차단한다. 디렉터리 1회 순회 + suffix
    # lower-case 비교로 ``.GIF``/``.PNG`` 같은 대문자 확장자도 포함한다.
    for entry in output_dir.iterdir():
        if not entry.is_file():
            continue
        ext = entry.suffix.lower()
        if ext != ".html" and ext not in _HTML_ASSET_SUFFIXES:
            continue
        try:
            entry.unlink()
        except OSError:
            pass

    last_error = ""
    for attempt in range(1, 4):
        profile_dir = Path(tempfile.mkdtemp(prefix="lo-profile-", dir="/tmp"))
        profile_uri = profile_dir.resolve().as_uri()
        try:
            args = [
                "--headless",
                "--invisible",
                "--nodefault",
                "--nolockcheck",
                "--nofirststartwizard",
                "--norestore",
                f"-env:UserInstallation={profile_uri}",
                "--convert-to",
                "html",
                "--outdir",
                str(output_dir),
                file_path,
            ]
            completed = _run_libreoffice_html_export(
                libreoffice_bin=libreoffice_bin,
                args=args,
                timeout=120,
            )
        finally:
            shutil.rmtree(profile_dir, ignore_errors=True)

        if completed.returncode == 0:
            break

        detail = (completed.stderr or completed.stdout or "").strip()
        last_error = detail or f"LibreOffice HTML export failed with code {completed.returncode}"
        log_info(f"[LibreOffice HTML Preview] 변환 실패 재시도 {attempt}/3: {last_error}")
        time.sleep(0.4 * attempt)
    else:
        raise RuntimeError(last_error or "LibreOffice HTML export failed")

    expected = output_dir / f"{Path(file_path).stem}.html"
    html_candidates = [expected] if expected.exists() else sorted(output_dir.glob("*.html"))
    source_html = next((candidate for candidate in html_candidates if candidate.exists()), None)
    if source_html is None:
        raise RuntimeError("LibreOffice 변환 결과 HTML을 찾을 수 없음")

    if source_html.resolve() != output_path.resolve():
        if output_path.exists():
            output_path.unlink()
        source_html.rename(output_path)


def _run_libreoffice_html_export(
    *,
    libreoffice_bin: str,
    args: list[str],
    timeout: int,
) -> subprocess.CompletedProcess[str]:
    """macOS 앱 등록 abort를 피하기 위해 직접 실행 실패 시 open 경로로 재시도한다."""

    direct = subprocess.run(
        [libreoffice_bin, *args],
        check=False,
        capture_output=True,
        text=True,
        timeout=timeout,
    )
    if direct.returncode == 0 or sys.platform != "darwin" or direct.returncode != -6:
        return direct

    # LibreOffice.app은 Codex/daemon 계열 프로세스에서 직접 실행하면 AppKit 등록 단계에서
    # Abort trap 6이 날 수 있다. open을 거치면 일반 앱 컨텍스트로 실행되어 변환이 완료된다.
    via_open = subprocess.run(
        ["open", "-W", "-n", "-a", "LibreOffice", "--args", *args],
        check=False,
        capture_output=True,
        text=True,
        timeout=timeout,
    )
    if via_open.returncode == 0:
        return via_open

    combined_stderr = "\n".join(
        part for part in (direct.stderr.strip(), via_open.stderr.strip()) if part
    )
    combined_stdout = "\n".join(
        part for part in (direct.stdout.strip(), via_open.stdout.strip()) if part
    )
    return subprocess.CompletedProcess(
        args=direct.args,
        returncode=via_open.returncode,
        stdout=combined_stdout,
        stderr=combined_stderr,
    )


def build_xlsx_html_preview_url(
    file_path: str,
    preview_output_dir: str,
    preview_base_url: str = "",
    *,
    job_token: str | None = None,
    subdir: str = "libreoffice-html",
    visible_sheets: int | None = None,
) -> str | None:
    """LibreOffice를 이용해 XLSX 원본 preview HTML을 생성하고 Azure SAS URL을 반환한다.

    정적 서빙은 제거됐다 — Azure 비활성 시 ``None`` 반환.
    """

    _plain_info(
        "[xlsx-preview] build start — file=%s subdir=%s job_token=%s visible_sheets=%s preview_dir=%s",
        file_path,
        subdir,
        job_token,
        visible_sheets,
        preview_output_dir,
    )

    if not preview_output_dir:
        _logger.warning(
            "[xlsx-preview] build skip — preview_output_dir 미설정 (AI_TRANSLATION_PREVIEW_ROOT 확인 필요)"
        )
        return None
    if not is_azure_preview_enabled():
        _logger.warning(
            "[xlsx-preview] build skip — Azure preview 비활성 (AZURE_STORAGE_CONNECTION_STRING 빈 값)"
        )
        return None

    job_id = job_token or uuid.uuid4().hex
    job_dir = os.path.join(preview_output_dir, job_id, subdir)
    html_path = Path(job_dir) / "index.html"
    source_file_path = file_path
    partial_workbook_path: Path | None = None
    try:
        if not job_token:
            cleanup_preview_output_dir(preview_output_dir)

        os.makedirs(job_dir, exist_ok=True)

        if isinstance(visible_sheets, int) and visible_sheets > 0:
            partial_workbook_path = _create_xlsx_visible_sheets_copy(file_path, visible_sheets)
            if partial_workbook_path is not None:
                source_file_path = str(partial_workbook_path)

        _export_office_html_with_libreoffice(source_file_path, str(html_path))
        _hide_xlsx_libreoffice_navigation(
            html_path,
            sheet_names=_read_xlsx_sheet_names(source_file_path),
        )

        url = _upload_html_with_assets(
            html_path,
            job_token=job_id,
            subdir=subdir,
        )
        if url is None:
            _logger.warning(
                "[xlsx-preview] build 실패 — _upload_html_with_assets 가 None 반환. html=%s",
                html_path,
            )
        else:
            _plain_info(
                "[xlsx-preview] build 성공 — job_id=%s subdir=%s html=%s",
                job_id,
                subdir,
                html_path,
            )
        return url
    except Exception as exc:
        _logger.warning(
            "[xlsx-preview] LibreOffice 변환 실패 — openpyxl fallback 시도: %s",
            exc,
            exc_info=True,
        )
        try:
            os.makedirs(job_dir, exist_ok=True)
            _write_xlsx_table_fallback_html(
                source_file_path if os.path.exists(source_file_path) else file_path,
                html_path,
                visible_sheets=visible_sheets,
            )
            url = _upload_html_with_assets(
                html_path,
                job_token=job_id,
                subdir=subdir,
            )
            if url is None:
                _logger.warning(
                    "[xlsx-preview] fallback build 실패 — _upload_html_with_assets 가 None. html=%s",
                    html_path,
                )
            else:
                _plain_info(
                    "[xlsx-preview] fallback build 성공 — job_id=%s subdir=%s html=%s",
                    job_id,
                    subdir,
                    html_path,
                )
            return url
        except Exception as fallback_exc:
            _logger.exception(
                "[xlsx-preview] fallback 도 실패: %s",
                fallback_exc,
            )
            return None
    finally:
        if partial_workbook_path is not None:
            try:
                partial_workbook_path.unlink()
                partial_workbook_path.parent.rmdir()
            except OSError:
                pass


def _create_xlsx_visible_sheets_copy(file_path: str, visible_sheets: int) -> Path | None:
    """완료된 시트까지만 남긴 preview 전용 임시 XLSX를 만든다."""

    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=False, keep_links=True)
    try:
        sheets = list(workbook.worksheets)
        if visible_sheets >= len(sheets):
            return None

        for sheet in sheets[visible_sheets:]:
            workbook.remove(sheet)
        if not workbook.worksheets:
            return None

        workbook.active = 0
        partial_dir = Path(tempfile.mkdtemp(prefix="xlsx-visible-sheets-", dir="/tmp"))
        partial_path = partial_dir / f"visible-{uuid.uuid4().hex}.xlsx"
        workbook.save(partial_path)
        return partial_path
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _write_xlsx_table_fallback_html(
    file_path: str,
    html_path: Path,
    *,
    visible_sheets: int | None = None,
) -> None:
    """LibreOffice가 실패한 XLSX에 대해 최소 HTML preview를 보장한다."""

    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell

    workbook = load_workbook(file_path, data_only=False)
    html_parts = [
        "<!doctype html>",
        "<html><head><meta charset='utf-8' />",
        "<style>",
        "body{margin:0;padding:24px;background:#f8fafc;color:#1f2638;font-family:Arial,'Apple SD Gothic Neo',sans-serif;}",
        ".sheet{margin:0 0 32px;padding:18px;border-radius:18px;background:white;box-shadow:0 10px 30px rgba(31,38,56,.08);}",
        ".sheet-title{font-size:18px;font-weight:700;margin:0 0 14px;color:#2563eb;}",
        "table{border-collapse:collapse;width:max-content;max-width:none;background:white;}",
        "td,th{border:1px solid #d7deea;padding:7px 10px;min-width:72px;max-width:360px;vertical-align:middle;white-space:pre-wrap;font-size:13px;line-height:1.35;}",
        "td:empty::after{content:' ';}",
        "</style></head><body>",
    ]

    worksheets = workbook.worksheets
    if isinstance(visible_sheets, int) and visible_sheets > 0:
        worksheets = worksheets[:visible_sheets]

    for sheet in worksheets:
        if sheet.title.strip().lower() == "evaluation warning":
            continue
        html_parts.append(
            f"<section class='sheet'><h2 class='sheet-title'>{escape(sheet.title)}</h2><table>"
        )
        merged_lookup: dict[tuple[int, int], tuple[int, int, bool]] = {}
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_lookup[(row, col)] = (
                        max_row - min_row + 1,
                        max_col - min_col + 1,
                        row == min_row and col == min_col,
                    )

        for row in sheet.iter_rows():
            html_parts.append("<tr>")
            for cell in row:
                if isinstance(cell, MergedCell):
                    info = merged_lookup.get((cell.row, cell.column))
                    if info and not info[2]:
                        continue
                rowspan, colspan, _ = merged_lookup.get((cell.row, cell.column), (1, 1, True))
                attrs = []
                if rowspan > 1:
                    attrs.append(f"rowspan='{rowspan}'")
                if colspan > 1:
                    attrs.append(f"colspan='{colspan}'")
                value = "" if cell.value is None else str(cell.value)
                html_parts.append(f"<td {' '.join(attrs)}>{escape(value)}</td>")
            html_parts.append("</tr>")
        html_parts.append("</table></section>")

    html_parts.append("</body></html>")
    html_path.write_text("\n".join(html_parts), encoding="utf-8")


def _read_xlsx_sheet_names(file_path: str) -> list[str]:
    """Preview 제목 복원에 사용할 XLSX 시트 이름을 읽는다."""

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(file_path, read_only=True, data_only=False)
        try:
            return [
                str(name)
                for name in workbook.sheetnames
                if str(name).strip().lower() != "evaluation warning"
            ]
        finally:
            workbook.close()
    except Exception:
        return []


def _hide_xlsx_libreoffice_navigation(html_path: Path, *, sheet_names: list[str] | None = None) -> None:
    """LibreOffice XLSX HTML의 Overview 링크를 숨기고 시트 제목은 preview에 남긴다."""

    try:
        html = html_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return

    html = re.sub(
        r"(<body[^>]*>)\s*<hr>\s*<p><center>\s*<h1>\s*Overview\s*</h1>.*?</center></p>\s*<hr>\s*",
        r"\1\n",
        html,
        count=1,
        flags=re.IGNORECASE | re.DOTALL,
    )

    title_index = 0

    def format_sheet_title(title: str) -> str:
        stripped = title.strip()
        if not stripped:
            return ""
        return stripped if stripped.endswith(":") else f"{stripped}:"

    def keep_sheet_title(match: re.Match[str]) -> str:
        nonlocal title_index
        anchor = match.group("anchor")
        raw_title = match.group("title")
        title = unescape(re.sub(r"<[^>]+>", "", raw_title)).strip()
        if not title and sheet_names and title_index < len(sheet_names):
            title = sheet_names[title_index]
        title_index += 1
        title = format_sheet_title(title)
        if not title:
            return f'<A NAME="{anchor}"></A>'
        return f'<A NAME="{anchor}"></A><div class="ai-xlsx-sheet-title">{escape(title)}</div>'

    html = re.sub(
        r'<A\s+NAME="(?P<anchor>table\d+)"\s*>\s*<h1>(?P<title>.*?)</h1>\s*</A>',
        keep_sheet_title,
        html,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if "ai-xlsx-sheet-title" not in html and sheet_names:
        title_index = 0

        def add_missing_sheet_title(match: re.Match[str]) -> str:
            nonlocal title_index
            anchor = match.group("anchor")
            title = ""
            if title_index < len(sheet_names):
                title = format_sheet_title(sheet_names[title_index])
            title_index += 1
            if not title:
                return match.group(0)
            return f'<A NAME="{anchor}"></A><div class="ai-xlsx-sheet-title">{escape(title)}</div>'

        html = re.sub(
            r'<A\s+NAME="(?P<anchor>table\d+)"\s*>\s*</A>',
            add_missing_sheet_title,
            html,
            flags=re.IGNORECASE,
        )

    injected_style = """
<style id="ai-xlsx-libreoffice-cleanup-style">
body { margin:0; padding:24px; background:#fff; }
a[name^="table"] > h1 {
  display:none !important;
}
a[name^="table"] {
  display:block;
  height:auto;
  overflow:hidden;
}
.ai-xlsx-sheet-title {
  margin:0 0 10px;
  color:#1f2638;
  font:700 18px/1.3 Arial, sans-serif;
}
table {
  margin-top:0 !important;
}
</style>
"""
    if "ai-xlsx-libreoffice-cleanup-style" in html:
        return
    if "</head>" in html:
        html = html.replace("</head>", f"{injected_style}</head>", 1)
    else:
        html = f"{injected_style}\n{html}"
    html_path.write_text(html, encoding="utf-8")


def _expand_xlsx_single_file_html(html_path: Path) -> None:
    """Aspose.Cells single-file HTML에서 모든 실제 sheet를 펼쳐 보이게 한다."""

    try:
        html = html_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return

    if "ai-xlsx-sheet-title" in html:
        return

    def replace_sheet_container(match: re.Match[str]) -> str:
        table_id = match.group("id")
        sheet_name = match.group("sheet")
        is_warning_sheet = sheet_name.strip().lower() == "evaluation warning"
        display = "none" if is_warning_sheet else "block"
        title = "" if is_warning_sheet else f"<div class='ai-xlsx-sheet-title'>{escape(sheet_name)}</div>"
        return f"<div id='table_{table_id}' style='display:{display}' sheetName='{sheet_name}'>{title}"

    html = re.sub(
        r"<div\s+id=['\"]table_(?P<id>\d+)['\"]\s+style=['\"]display:(?:none|block)['\"]\s+sheetName=['\"](?P<sheet>[^'\"]+)['\"]>",
        replace_sheet_container,
        html,
        flags=re.IGNORECASE,
    )

    injected_style = """
<style id="ai-xlsx-preview-style">
html, body { height:auto !important; overflow:auto !important; }
#section {
  height:auto !important;
  width:auto !important;
  overflow:visible !important;
  float:none !important;
  padding:16px !important;
}
#footer { display:none !important; }
.ai-xlsx-sheet-title {
  margin:18px 0 10px;
  padding:8px 12px;
  border-radius:10px;
  background:#f1f5fb;
  color:#1f2638;
  font:600 14px/1.3 Arial, sans-serif;
}
div[sheetName="Evaluation Warning"] { display:none !important; }
</style>
"""
    if "</head>" in html:
        html = html.replace("</head>", f"{injected_style}</head>", 1)
    else:
        html = f"{injected_style}\n{html}"

    html_path.write_text(html, encoding="utf-8")


def _truncate_xlsx_html_to_visible_sheets(html_path: Path, visible_sheets: int) -> None:
    """LibreOffice XLSX HTML에서 완료된 시트까지만 남긴다."""

    if visible_sheets <= 0:
        return

    try:
        html = html_path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return

    sheet_blocks = _find_libreoffice_xlsx_sheet_blocks(html)
    if len(sheet_blocks) <= visible_sheets:
        return

    kept_blocks = sheet_blocks[:visible_sheets]
    first_sheet_start = sheet_blocks[0][0]
    last_kept_end = kept_blocks[-1][1]
    body_tail_start = sheet_blocks[-1][1]

    truncated = html[:first_sheet_start] + html[first_sheet_start:last_kept_end] + html[body_tail_start:]

    def remove_hidden_overview_link(match: re.Match[str]) -> str:
        table_index = int(match.group("index"))
        return match.group(0) if table_index < visible_sheets else ""

    truncated = re.sub(
        r'\s*<A\s+HREF="#table(?P<index>\d+)">.*?</A><br>\s*',
        remove_hidden_overview_link,
        truncated,
        flags=re.IGNORECASE | re.DOTALL,
    )

    injected_style = """
<style id="ai-xlsx-visible-sheets-style">
body { margin:0; padding:24px; background:#f8fafc; color:#1f2638; }
hr:first-of-type, hr:nth-of-type(2) { display:none; }
body > p:first-of-type { display:none; }
a[name^="table"] h1, a[name^="table"] + h1 {
  margin:0 0 16px;
  font:700 18px/1.3 Arial, "Apple SD Gothic Neo", sans-serif;
  color:#2563eb;
}
table {
  margin:0 0 32px;
  background:white;
  box-shadow:0 10px 30px rgba(31,38,56,.08);
}
</style>
"""
    if "</head>" in truncated:
        truncated = truncated.replace("</head>", f"{injected_style}</head>", 1)
    else:
        truncated = f"{injected_style}\n{truncated}"

    html_path.write_text(truncated, encoding="utf-8")


def _find_libreoffice_xlsx_sheet_blocks(html: str) -> list[tuple[int, int]]:
    """LibreOffice XLSX HTML의 각 시트 anchor/table 블록 범위를 찾는다."""

    starts = [match.start() for match in re.finditer(r'<A\s+NAME="table\d+"\s*>', html, flags=re.IGNORECASE)]
    if not starts:
        return []

    body_close = html.lower().rfind("</body>")
    if body_close < 0:
        body_close = len(html)

    blocks: list[tuple[int, int]] = []
    for index, start in enumerate(starts):
        end = starts[index + 1] if index + 1 < len(starts) else body_close
        blocks.append((start, end))
    return blocks


def append_preview_version(url: str | None, version: str | int | None) -> str | None:
    """브라우저 캐시를 우회하기 위한 version query를 preview URL에 덧붙인다."""

    if not url or version is None:
        return url

    parts = urlsplit(url)
    query = parts.query
    version_query = f"v={version}"
    query = f"{query}&{version_query}" if query else version_query
    return urlunsplit((parts.scheme, parts.netloc, parts.path, query, parts.fragment))


_DIV_TAG_RE = re.compile(r"<div\b[^>]*>|</div\s*>", re.IGNORECASE)
_SLIDE_CLASS_RE = re.compile(r'\bclass\s*=\s*["\'][^"\']*\bslide\b[^"\']*["\']', re.IGNORECASE)


def _truncate_pptx_html_to_visible_slides(html_path: str, visible_slides: int) -> None:
    """Aspose HTML 결과에서 지정한 수만큼의 slide div만 남긴다."""

    if visible_slides <= 0:
        return

    path = Path(html_path)
    html = path.read_text(encoding="utf-8", errors="ignore")
    slide_blocks = _find_top_level_slide_blocks(html)
    if len(slide_blocks) <= visible_slides:
        return

    kept_blocks = slide_blocks[:visible_slides]
    new_html = (
        html[: slide_blocks[0][0]]
        + "".join(html[start:end] for start, end in kept_blocks)
        + html[slide_blocks[-1][1] :]
    )
    path.write_text(new_html, encoding="utf-8")


def _find_top_level_slide_blocks(html: str) -> list[tuple[int, int]]:
    """Aspose HTML에서 최상위 slide div 블록 범위를 찾는다."""

    stack: list[tuple[int, bool]] = []
    blocks: list[tuple[int, int]] = []

    for match in _DIV_TAG_RE.finditer(html):
        tag = match.group(0)
        is_close = tag.lower().startswith("</div")

        if is_close:
            if not stack:
                continue
            start, is_slide = stack.pop()
            if is_slide and not any(parent_is_slide for _, parent_is_slide in stack):
                blocks.append((start, match.end()))
            continue

        is_slide = bool(_SLIDE_CLASS_RE.search(tag))
        stack.append((match.start(), is_slide))

    return blocks


def build_office_preview_result(
    file_path: str,
    nodes: list[dict],
    ext: str,
    preview_output_dir: str,
    preview_base_url: str,
    deps: OfficePipelineDeps,
) -> PreviewPayload:
    """Office 문서 preview payload를 생성하고 필요 시 외부 URL로 치환한다.

    Args:
        file_path: preview 생성을 위한 문서 경로.
        nodes: 번역 결과가 반영된 노드 목록.
        ext: 파일 확장자.
        preview_output_dir: preview 파일 저장 디렉터리.
        preview_base_url: preview 파일 접근 base URL.
        deps: preview 단계에서 필요한 의존성 묶음.

    Returns:
        프런트에서 사용하는 preview payload.
    """

    preview_payload = deps.build_office_preview_payload(file_path, nodes, ext)
    return deps.externalize_preview_payload(
        preview_payload,
        preview_output_dir,
        preview_base_url,
    )


def count_preview_payload_chars(preview_payload: PreviewPayload) -> int:
    """preview payload 내부 문자열 길이를 합산한다.

    Args:
        preview_payload: preview 이미지/URL 목록을 담은 payload.

    Returns:
        문자열 총 길이 합계.
    """

    return sum(
        len(item)
        for key in ("original_preview_images", "translated_preview_images")
        for item in preview_payload.get(key, [])
        if isinstance(item, str)
    )
